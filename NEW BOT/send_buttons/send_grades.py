import io
from datetime import datetime, timedelta
import openpyxl
from telegram import Update
from telegram.ext import ContextTypes, ConversationHandler
from database import get_connection, insert_grade, get_room_id_by_number
from config import ENTERING_DATE, WAITING_EXCEL_FILE, CATEGORY_ROWS

async def info_excel_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Excel yuklash uchun dastlab sanani kiriting (YYYY-MM-DD):")
    return ENTERING_DATE

async def date_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    date_str = update.message.text.strip()
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        context.user_data["upload_date"] = date_str
        await update.message.reply_text(f"Sana {date_str} deb qabul qilindi. Endi Excel faylni yuboring:")
        return WAITING_EXCEL_FILE
    except ValueError:
        await update.message.reply_text("Sana formati noto'g'ri. Qaytadan kiriting (YYYY-MM-DD):")
        return ENTERING_DATE

async def excel_file_received(update: Update, context: ContextTypes.DEFAULT_TYPE):
    doc = update.message.document
    if not (doc.file_name.endswith(".xlsx") or doc.file_name.endswith(".xls")):
        await update.message.reply_text("Faqat Excel fayl yuboring!")
        return WAITING_EXCEL_FILE
    
    file_bytes = await context.bot.get_file(doc.file_id)
    f_io = io.BytesIO()
    await file_bytes.download_to_memory(out=f_io)
    f_io.seek(0)

    wb = openpyxl.load_workbook(f_io, data_only=True)
    sheet = wb.active
    date_str = context.user_data.get("upload_date")

    for col in range(2, sheet.max_column + 1):
        room_num = sheet.cell(row=1, column=col).value
        if not room_num:
            continue
        room_id = get_room_id_by_number(str(room_num))
        if not room_id:
            continue
        for r_idx, cat_name in CATEGORY_ROWS.items():
            val = sheet.cell(row=r_idx, column=col).value
            if val is not None:
                insert_grade(room_id, date_str, cat_name, int(val))

    await update.message.reply_text("Baholar muvaffaqiyatli bazaga yozildi!")
    return ConversationHandler.END